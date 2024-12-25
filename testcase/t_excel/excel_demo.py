from openpyxl import load_workbook

def read_excel_file():
    workbook = load_workbook('api tc.xlsx')

    # 通过名称获取sheet
    sheet = workbook['Sheet1']
    # 获取行数和列数
    rows = sheet.max_row
    cols = sheet.max_column

    # 遍历每一行
    for row in sheet.iter_rows(min_row=2, min_col=1, max_row=rows, max_col=cols, values_only=True):
        # 遍历每一列
        for cell in row:
            # 打印每一个单元格的值
            print(cell)

    # 获取某一列的信息
    col_values = [cell.value for cell in sheet['A']]
    print(col_values)

if __name__ == '__main__':
    read_excel_file()
